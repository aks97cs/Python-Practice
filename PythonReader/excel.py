import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
# print(type(wb))
# sheet_name = wb.get_sheet_names()
sheet_name = wb.sheetnames
print("*********** Workbook Information **************")
for i in sheet_name:
	print(i,end=" ")
select_sheet = wb['Sheet1']
# select_sheet = wb.get_sheet_by_name('Sheet1')

# print("select sheet : ",select_sheet)
print("")
print("********** Selected sheet is : ***********")
print("selected sheet name  : ",select_sheet.title)
# count = 0
print("************** Excel File Data *****************")
# for i in select_sheet:
# 	for i in select_sheet:
# 		print(i[count].value)
# 	count+=1

# for i in select_sheet:
# 	print(i)


# print(select_sheet.cell(row=1,column=2))
# print(select_sheet.cell(row=1,column=2).value)
# print(select_sheet.get_highest_row())
# print(dir(select_sheet)) # give list of attr that we could use
# print(select_sheet.max_column) # give number of col
# print(select_sheet.get_highest_column())

row = select_sheet.max_row
col = select_sheet.max_column
for i in range(1,row+1):
	for j in range(1,col+1):
		print(select_sheet.cell(i,j).value,end = "|")
	print("")

print("*********** Updating Row ****************")
for i in range(1,row+1):
	for j in range(1,col+1):
		# print(select_sheet.cell(i,j).value,end = "|")
		print(type(select_sheet[i][j]))
	print("row",i,"has been updated")


# print(select_sheet['A2'].value)

# print(select_sheet)

# print(select_sheet['A1'])
# print(select_sheet['A1'].value)

# select_active_sheet = wb.get_active_sheet()

# print("Current sheel is |",select_active_sheet)